from tkinter import *
from openpyxl import *

wb = load_workbook("Book1.xlsx")

sheet = wb.active



def excel():
    sheet.cell(row=1, column=1).value = "First Name"
    sheet.cell(row=1, column=2).value = "Last Name"
    sheet.cell(row=1, column=3).value = "Username"
    sheet.cell(row=1, column=4).value = "Password"
    sheet.cell(row=1, column=5).value = "Email"
    sheet.cell(row=1, column=6).value = "Gender"
    sheet.cell(row=1, column=7).value = "Age"
    sheet.cell(row=1, column=8).value = "Agree"


def insert():
    current_row = sheet.max_row

    sheet.cell(row=current_row + 1, column=1).value = firstname.get()
    sheet.cell(row=current_row + 1, column=2).value = lastname.get()
    sheet.cell(row=current_row + 1, column=3).value = username.get()
    sheet.cell(row=current_row + 1, column=4).value = password.get()
    sheet.cell(row=current_row + 1, column=5).value = email.get()
    sheet.cell(row=current_row + 1, column=7).value = age.get()
    sheet.cell(row=current_row + 1, column=6).value = gend.get()
    sheet.cell(row=current_row + 1, column=8).value = agree.get()

    wb.save("Book1.xlsx")

    firstname.focus_set()

excel()
root = Tk()


root.configure(background = "blue")
root.title("Excel")


Label(root, text="First Name").grid(row = 0, column =0 )
Label(root, text="Last Name").grid(row = 1, column = 0)
Label(root, text="Username").grid(row = 2, column = 0)
Label(root, text="Password").grid(row = 3, column = 0)
Label(root, text="Email").grid(row = 4, column = 0)
Label(root, text="Age").grid(row = 5, column = 0)

firstname = Entry(root)
firstname.grid(row = 0, column = 1)
lastname = Entry(root)
lastname.grid(row = 1, column = 1)
username =Entry(root)
username.grid(row = 2, column = 1)
password = Entry(root)
password.grid(row = 3, column = 1)
email = Entry(root)
email.grid(row = 4, column = 1)
age = Scale(root, from_=0, to = 100, orient = HORIZONTAL)
age.grid(row = 5, column = 1)
gend = StringVar()
Radiobutton(root, text = "Male", variable = gend, value = "male").grid(row = 6, column = 0)
Radiobutton(root, text = "Female", variable = gend, value = "female").grid(row = 6, column = 1)
agree = IntVar()
Checkbutton(root, variable = agree, text = "Accept policy").grid(row = 7, column = 0)

Button(root, text = "Insert", command = insert).grid(row = 8, column = 0)
root.mainloop()